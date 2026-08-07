"""
Genera sugerencias de Listas de Materiales (BoM) para Odoo.

Lógica de detección:
- Unidades (1U, 3U, 6U): el pack más chico es componente de los más grandes.
- Peso (G, K): si hay GRL, todos los fraccionados vienen del granel.
  Si no hay GRL, el fraccionado más grande es componente del siguiente menor.
- Volumen (ML, CC, L): igual que peso.
- Sabores (NAT, VAI, FRRJ, etc.): no generan BoM, son productos independientes.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database.models import Producto, SKU


# ── Parsing de presentaciones ────────────────────────────────────────────────

def _parse_pres(codigo):
    """
    Retorna (valor_en_unidad_base, tipo) o None si no aplica.
    Tipos: 'unit', 'weight', 'volume'
    Unidades base: unidades, gramos, mililitros
    """
    c = codigo.upper()

    # Granel → peso infinito como fuente
    if c == "GRL":
        return (float("inf"), "weight")

    # Unidades: 1U, 3U, 6U, 10U ...
    m = re.match(r"^(\d+)U$", c)
    if m:
        return (int(m.group(1)), "unit")

    # Sets (1 unidad compuesta)
    if c == "SET":
        return (1, "unit")

    # Peso en gramos: 50G, 100G, 250G, 500G, 800G / trozos (T) / pulpa (P)
    m = re.match(r"^(\d+)[GTP]$", c)
    if m:
        return (int(m.group(1)), "weight")

    # Peso en kilos: 1K, 2K, 5K, 10K, 15K, 25K, 30K
    m = re.match(r"^(\d+)K$", c)
    if m:
        return (int(m.group(1)) * 1000, "weight")

    # 1K5 → 1500g
    m = re.match(r"^(\d+)K(\d+)$", c)
    if m:
        return (int(m.group(1)) * 1000 + int(m.group(2)) * 100, "weight")

    # Volumen en ml: 500ML
    m = re.match(r"^(\d+)ML$", c)
    if m:
        return (int(m.group(1)), "volume")

    # Volumen en cc: 200C, 360C
    m = re.match(r"^(\d+)C$", c)
    if m:
        return (int(m.group(1)), "volume")

    # Volumen en litros: 1L, 2L
    m = re.match(r"^(\d+)L$", c)
    if m:
        return (int(m.group(1)) * 1000, "volume")

    return None


def _es_sabor(codigo):
    SABORES = {"NAT", "VAI", "FRRJ", "LECH", "BLCK", "BLAN"}
    return codigo.upper() in SABORES


def _parse_pack_pres(codigo):
    """
    Detecta packs tipo 6X250 → (6, 250): 6 unidades de la presentación de 250ml/g.
    Retorna (count, base_size) o None.
    """
    m = re.match(r"^(\d+)X(\d+)$", codigo.upper())
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


# ── Lógica de sugerencias ────────────────────────────────────────────────────

def _sugerir_boms_producto(producto):
    """
    Analiza los SKUs de un producto y retorna lista de dicts:
    {producto_final, sku_final, componente, sku_componente, cantidad, tipo, notas}
    """
    skus = producto.skus
    if len(skus) < 2:
        return []

    # Separar por tipo
    grupos = {}        # tipo → [(valor, sku_obj)]
    sabores = []
    pack_skus = []     # (count, base_size, sku_obj) para packs nXm

    for s in skus:
        if _es_sabor(s.presentacion):
            sabores.append(s)
            continue
        pack = _parse_pack_pres(s.presentacion)
        if pack:
            count, base_size = pack
            pack_skus.append((count, base_size, s))
            continue
        parsed = _parse_pres(s.presentacion)
        if parsed is None:
            continue
        valor, tipo = parsed
        grupos.setdefault(tipo, []).append((valor, s))

    boms = []

    for tipo, items in grupos.items():
        if len(items) < 2:
            continue

        items_sorted = sorted(items, key=lambda x: x[0])  # menor a mayor
        tipo_label = {"unit": "Pack de unidades", "weight": "Fraccionado", "volume": "Fraccionado"}[tipo]

        tiene_granel = any(v == float("inf") for v, _ in items_sorted)

        if tipo == "unit":
            # El más chico es la unidad base; los demás son packs de ese base
            base_val, base_sku = items_sorted[0]
            for val, sku in items_sorted[1:]:
                if base_val > 0 and val % base_val == 0:
                    cantidad = int(val / base_val)
                    boms.append({
                        "producto_final": f"{producto.nombre} ({sku.presentacion})",
                        "sku_final": sku.sku,
                        "componente": f"{producto.nombre} ({base_sku.presentacion})",
                        "sku_componente": base_sku.sku,
                        "cantidad": cantidad,
                        "tipo": tipo_label,
                        "notas": "",
                    })
                else:
                    boms.append({
                        "producto_final": f"{producto.nombre} ({sku.presentacion})",
                        "sku_final": sku.sku,
                        "componente": f"{producto.nombre} ({base_sku.presentacion})",
                        "sku_componente": base_sku.sku,
                        "cantidad": f"{val}/{base_val} — verificar",
                        "tipo": tipo_label,
                        "notas": "Relación no exacta, revisar manualmente",
                    })

        elif tipo in ("weight", "volume"):
            if tiene_granel:
                granel_sku = next(s for v, s in items_sorted if v == float("inf"))
                fraccionados = [(v, s) for v, s in items_sorted if v != float("inf")]
                for val, sku in fraccionados:
                    unidad = "g" if tipo == "weight" else "ml"
                    boms.append({
                        "producto_final": f"{producto.nombre} ({sku.presentacion})",
                        "sku_final": sku.sku,
                        "componente": f"{producto.nombre} ({granel_sku.presentacion})",
                        "sku_componente": granel_sku.sku,
                        "cantidad": f"{val}{unidad}",
                        "tipo": tipo_label,
                        "notas": "Fraccionado desde granel",
                    })
            else:
                # Sin granel: el más grande es fuente del siguiente más chico
                for i in range(len(items_sorted) - 1, 0, -1):
                    val_fin, sku_fin = items_sorted[i - 1]   # más chico = producto final
                    val_comp, sku_comp = items_sorted[i]     # más grande = componente/fuente
                    if val_fin == val_comp:
                        continue  # mismo tamaño pero distinto tipo (ej: trozos vs pulpa)
                    if val_fin > 0 and val_comp % val_fin == 0:
                        cantidad = int(val_comp / val_fin)
                        boms.append({
                            "producto_final": f"{producto.nombre} ({sku_fin.presentacion})",
                            "sku_final": sku_fin.sku,
                            "componente": f"{producto.nombre} ({sku_comp.presentacion})",
                            "sku_componente": sku_comp.sku,
                            "cantidad": cantidad,
                            "tipo": tipo_label,
                            "notas": "",
                        })
                    else:
                        boms.append({
                            "producto_final": f"{producto.nombre} ({sku_fin.presentacion})",
                            "sku_final": sku_fin.sku,
                            "componente": f"{producto.nombre} ({sku_comp.presentacion})",
                            "sku_componente": sku_comp.sku,
                            "cantidad": "— verificar",
                            "tipo": tipo_label,
                            "notas": "Relación no múltiplo exacto, revisar",
                        })

    # Packs nXm: 6X250 → BoM: 6× la presentación de 250ml/g
    if pack_skus:
        size_by_val = {}
        for tipo_k, items_k in grupos.items():
            if tipo_k in ("volume", "weight"):
                for v, s in items_k:
                    size_by_val[v] = s
        for count, base_size, pack_sku in pack_skus:
            base_sku = size_by_val.get(base_size)
            if base_sku:
                boms.append({
                    "producto_final": f"{producto.nombre} ({pack_sku.presentacion})",
                    "sku_final": pack_sku.sku,
                    "componente": f"{producto.nombre} ({base_sku.presentacion})",
                    "sku_componente": base_sku.sku,
                    "cantidad": count,
                    "tipo": "Pack de unidades",
                    "notas": "",
                })
            else:
                # Fallback: usar granel con peso total si existe
                granel_sku = next(
                    (s for v, s in grupos.get("weight", []) if v == float("inf")),
                    None,
                )
                if granel_sku:
                    total_g = count * base_size
                    boms.append({
                        "producto_final": f"{producto.nombre} ({pack_sku.presentacion})",
                        "sku_final": pack_sku.sku,
                        "componente": f"{producto.nombre} ({granel_sku.presentacion})",
                        "sku_componente": granel_sku.sku,
                        "cantidad": f"{total_g}g",
                        "tipo": "Pack de unidades",
                        "notas": "",
                    })
                else:
                    boms.append({
                        "producto_final": f"{producto.nombre} ({pack_sku.presentacion})",
                        "sku_final": pack_sku.sku,
                        "componente": f"— presentación de {base_size} no encontrada",
                        "sku_componente": "",
                        "cantidad": count,
                        "tipo": "Pack de unidades",
                        "notas": f"No hay presentación de {base_size}ml/g en la DB — verificar manualmente",
                    })

    return boms


# ── Exportación ──────────────────────────────────────────────────────────────

def exportar_boms(session, archivo_path):
    productos = session.query(Producto).order_by(Producto.sku_base).all()

    todas_boms = []
    for p in productos:
        todas_boms.extend(_sugerir_boms_producto(p))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BoMs sugeridas"

    # Encabezados
    headers = [
        "Producto Final", "SKU Final",
        "Componente (fuente)", "SKU Componente",
        "Cantidad", "Tipo", "Notas",
    ]
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas
    warn_fill = PatternFill("solid", fgColor="FFF9C4")  # amarillo suave para "verificar"

    for row_idx, bom in enumerate(todas_boms, 2):
        valores = [
            bom["producto_final"],
            bom["sku_final"],
            bom["componente"],
            bom["sku_componente"],
            bom["cantidad"],
            bom["tipo"],
            bom["notas"],
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val != "" else "")
            if bom["notas"]:
                cell.fill = warn_fill

    # Anchos de columna
    anchos = [45, 18, 45, 18, 12, 22, 35]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho

    ws.freeze_panes = "A2"

    # Segunda hoja: productos sin BoM sugerida (solo 1 presentación)
    ws2 = wb.create_sheet("Sin BoM (1 presentación)")
    ws2.append(["Producto", "SKU Base", "Categoría", "Presentación única"])
    h2_fill = PatternFill("solid", fgColor="1565C0")
    h2_font = Font(bold=True, color="FFFFFF")
    for cell in ws2[1]:
        cell.fill = h2_fill
        cell.font = h2_font

    productos_con_bom = {b["sku_final"].rsplit("-", 1)[0] for b in todas_boms}
    productos_con_bom |= {b["sku_componente"].rsplit("-", 1)[0] for b in todas_boms}

    for p in productos:
        if p.sku_base not in productos_con_bom:
            for s in p.skus:
                ws2.append([p.nombre, p.sku_base, p.categoria.codigo, s.presentacion])

    for col, ancho in zip("ABCD", [45, 12, 12, 16]):
        ws2.column_dimensions[col].width = ancho

    wb.save(archivo_path)
    return len(todas_boms)
